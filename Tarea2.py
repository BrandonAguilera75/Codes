#Aguilera Salinas Brandon Alfonso
import pandas as pd 
import numpy as np
import os 
import csv
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import  PatternFill, Font
from openpyxl.chart import  Reference, BarChart
from openpyxl.utils import get_column_letter
import xlwings as xw
from openpyxl.styles import (NamedStyle, colors, builtins, Alignment, Border, borders, Side, Font, fonts, fills, PatternFill, numbers)

def llenar(x,y):
    for elem in df[y]:
        x.append(elem)
    return x

def chantype(x,y):
    a=pd.DataFrame(x).astype(y)
    return a

def maxLength(col):
    max_length = -math.inf
    for cell in col:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    return (max_length + 2)


def Save(fileName):
   absolute_path=os.path.dirname(__file__)
   return os.path.join(absolute_path, fileName)


if __name__ == "__main__" :
    wb = Workbook()
    ws=wb.active
    
    relativepath= os.path.dirname(__file__)
    path_to_open="muestra_de_base_comext.tsv"
    full_path=os.path.join(relativepath, path_to_open)
    df=pd.read_csv(full_path, encoding="utf8", delimiter="\t")

    
    
    estatus=[]
    anio=[]
    mes=[]
    tipo=[]
    fracc=[]
    umed=[]
    pais=[]
    aduana=[]
    mtra=[]
    cantidad=[]
    valmx=[]
    valusd=[]
    estatus=llenar(estatus,"ESTATUS")
    anio=llenar(anio,"ANIO")
    mes=llenar(mes,"MES")
    tipo=llenar(tipo,"TIPO")
    fracc=llenar(fracc,"FRACCION")
    umed=llenar(umed,"UMED")
    pais=llenar(pais,"PAIS_O_D")
    aduana=llenar(aduana,"ADUANA")
    mtra=llenar(mtra,"MTRA")
    cantidad=llenar(cantidad, "CANTIDAD")
    valmx=llenar(valmx, "VAL_MNX")
    valusd=llenar(valusd, "VAL_USD")
    
    
    
    df["ESTATUS"]=chantype(estatus, "object")
    df["ANIO"]=chantype(anio, "int64")
    df["MES"]=chantype(mes, "int64")
    df["TIPO"]=chantype(tipo, "object")
    df["FRACCION"]=chantype(fracc, "int64")
    df["UMED"]=chantype(umed, "object")
    df["PAIS_O_D"]=chantype(pais, "object")
    df["ADUANA"]=chantype(aduana, "int64")
    df["MTRA"]=chantype(mtra, "int64")
    df["CANTIDAD"]=chantype(cantidad, "float64")
    df["VAL_MNX"]=chantype(valmx, "float64")
    df["VAL_USD"]=chantype(valusd, "float64")


    df.to_excel('data_tarea2.xlsx', index=False) #excell sin indice
    
    lista=df.values.tolist()
    
    Header = ["ESTATUS","AÑO","MES","FRACCION","UMED","PAIS_O_D","ADUANA","MTRA","CANTIDAD","VAL_MNX","VAL_USD"]
    ws.append(Header)
    
    for row in lista: 
        ws.append(row)

    for row in ws["A1:L1"]:
        for cell in row:
         cell.fill=PatternFill(start_color="FFAAAA", fill_type="solid")
         cell.font=Font(bold=True)
    
    for row in ws["A2:L1001"]:
        for cell in row:
         cell.fill=PatternFill(start_color="AAFFAA", fill_type="solid")
         cell.font=Font(bold=True)

    for col in ws.columns:
        max_length = -math.inf
        column = col[0].column_letter 
        ws.column_dimensions[column].width = maxLength(col)    

    promedio=np.mean(cantidad)

    Fecha=[f"{anio}-{mes}" for anio, mes in zip(anio,mes) ]
    
    Fecha=chantype(Fecha, "datetime64[ns]")
    df["ANIO"]=Fecha
    df=df.drop(columns="MES")


    wb.save(Save("data_tarea.xlsx"))
    print(promedio)
    print(df)





    
    
    
    

    


