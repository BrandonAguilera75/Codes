import pandas as pd 
import os 
import csv
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import  PatternFill, Font
from openpyxl.chart import  Reference, BarChart

def Save(fileName):
   absolute_path=os.path.dirname(__file__)
   return os.path.join(absolute_path, fileName)

def longmax(arr):
   max=-math.inf
   for elem in arr:
      if len(elem)>max:
         max=len(elem)
   return max

   
if __name__ == "__main__" :
    wb = Workbook()
    ws=wb.active

    treeData=[
        ["Type", "Leaf Color","Heigth" ],
        ["Maple","Red",549],
        ["Oak","Green",783],
        ["Pine","Green",1204]
    ]
    
    for row in treeData:
        ws.append(row)
    
    for row in ws["A1:C1"]:
        for cell in row:
         cell.fill=PatternFill(start_color="FFAAAA", fill_type="solid")
         cell.font=Font(bold=True)

    col= ws.column_dimensions["B"]
    col.width=20
    
    chart= BarChart()
    chart.type="col"

    chart.title="Tree Heigth"
    chart.x_axis.title="Tree Type"
    chart.y_axis.title="Heigth(cm)"
    chart.legend=None

    data=Reference(ws, min_col=3, max_col=3, min_row=2, max_row=4)
    categories=Reference(ws, min_col=1, max_col=1, min_row=2, max_row=4)

    chart.add_data(data)
    chart.set_categories(categories)

    ws.add_chart(chart, "E1")
    wb.save(Save("muestra_formato.xlsx"))

